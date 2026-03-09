import pandas as pd
import numpy as np
from io import BytesIO
from datetime import datetime
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from fpdf import FPDF
import tempfile
import os

class ExcelExporter:
    """Class untuk export ke Excel"""
    
    @staticmethod
    def export_to_excel(dataframe, sheet_name="Hasil Scanning"):
        """Export DataFrame ke Excel dengan formatting"""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            dataframe.to_excel(writer, sheet_name=sheet_name, index=False)
            
            # Get workbook and worksheet
            workbook = writer.book
            worksheet = writer.sheets[sheet_name]
            
            # Styling
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Apply header styling
            for col in range(1, len(dataframe.columns) + 1):
                cell = worksheet.cell(row=1, column=col)
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = center_alignment
            
            # Auto-adjust column widths
            for col in worksheet.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                worksheet.column_dimensions[column].width = adjusted_width
            
            # Add border to all cells
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            for row in worksheet.iter_rows(min_row=1, max_row=len(dataframe) + 1, min_col=1, max_col=len(dataframe.columns)):
                for cell in row:
                    cell.border = thin_border
        
        output.seek(0)
        return output.getvalue()
    
    @staticmethod
    def export_multiple_sheets(dataframes_dict):
        """Export multiple DataFrames ke different sheets"""
        output = BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, df in dataframes_dict.items():
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)  # Sheet name max 31 chars
        
        output.seek(0)
        return output.getvalue()

class PDFExporter(FPDF):
    """Class untuk export ke PDF"""
    
    def __init__(self):
        super().__init__()
        self.set_auto_page_break(auto=True, margin=15)
    
    def header(self):
        """Header setiap halaman"""
        self.set_font('Arial', 'B', 12)
        self.cell(0, 10, 'Hasil Scanning Saham Indonesia', 0, 1, 'C')
        self.set_font('Arial', 'I', 8)
        self.cell(0, 5, f'Generated: {datetime.now().strftime("%Y-%m-%d %H:%M:%S")}', 0, 1, 'C')
        self.ln(5)
    
    def footer(self):
        """Footer setiap halaman"""
        self.set_y(-15)
        self.set_font('Arial', 'I', 8)
        self.cell(0, 10, f'Halaman {self.page_no()}', 0, 0, 'C')
    
    def add_title(self, title):
        """Menambahkan judul section"""
        self.set_font('Arial', 'B', 11)
        self.set_fill_color(200, 220, 255)
        self.cell(0, 8, title, 0, 1, 'L', 1)
        self.ln(2)
    
    def add_summary(self, summary_dict):
        """Menambahkan ringkasan"""
        self.set_font('Arial', '', 10)
        for key, value in summary_dict.items():
            self.cell(50, 6, f"{key}:", 0, 0)
            self.cell(0, 6, str(value), 0, 1)
        self.ln(5)
    
    def add_table(self, dataframe, max_rows=20):
        """Menambahkan tabel dari DataFrame"""
        if dataframe.empty:
            self.cell(0, 6, "Tidak ada data", 0, 1)
            return
        
        # Limit rows untuk PDF
        df_display = dataframe.head(max_rows)
        
        # Column widths (adjust based on content)
        col_widths = []
        for col in df_display.columns:
            max_len = max(
                df_display[col].astype(str).str.len().max(),
                len(str(col))
            )
            width = min(max_len * 2, 40)  # Max 40mm
            col_widths.append(width)
        
        # Table header
        self.set_font('Arial', 'B', 9)
        self.set_fill_color(200, 200, 200)
        for i, col in enumerate(df_display.columns):
            self.cell(col_widths[i], 7, str(col), 1, 0, 'C', 1)
        self.ln()
        
        # Table data
        self.set_font('Arial', '', 8)
        fill = False
        for _, row in df_display.iterrows():
            for i, value in enumerate(row):
                # Format numbers
                if isinstance(value, float):
                    text = f"{value:.2f}"
                else:
                    text = str(value)
                
                self.cell(col_widths[i], 6, text[:15], 1, 0, 'L', fill)
            self.ln()
            fill = not fill
        
        if len(dataframe) > max_rows:
            self.cell(0, 6, f"... dan {len(dataframe) - max_rows} data lainnya", 0, 1)
        
        self.ln(5)

def export_to_excel(dataframe, sheet_name="Hasil Scanning"):
    """Wrapper function untuk export Excel"""
    exporter = ExcelExporter()
    return exporter.export_to_excel(dataframe, sheet_name)

def export_to_pdf(dataframe, title="Hasil Scanning", summary=None):
    """Wrapper function untuk export PDF"""
    pdf = PDFExporter()
    pdf.add_page()
    
    # Add title
    pdf.add_title(title)
    
    # Add summary if provided
    if summary:
        pdf.add_summary(summary)
    
    # Add table
    pdf.add_table(dataframe)
    
    # Output to BytesIO
    output = BytesIO()
    pdf.output(output, 'F')  # Write to BytesIO
    
    # Baca file yang ditulis
    with open(output.name, 'rb') as f:
        pdf_data = f.read()
    
    # Clean up temp file
    os.unlink(output.name)
    
    return pdf_data

def export_to_csv(dataframe):
    """Export ke CSV"""
    output = BytesIO()
    dataframe.to_csv(output, index=False)
    output.seek(0)
    return output.getvalue()

def format_number(num):
    """Format angka untuk display"""
    if num >= 1e9:
        return f"{num/1e9:.2f} B"
    elif num >= 1e6:
        return f"{num/1e6:.2f} M"
    elif num >= 1e3:
        return f"{num/1e3:.2f} K"
    else:
        return str(num)
